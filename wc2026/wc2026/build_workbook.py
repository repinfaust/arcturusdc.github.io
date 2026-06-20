"""Build the WC2026 value-finder workbook from the engine output."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from engine import build_match
from markets import market_1x2
from value_finder import load_ratings, find_value

FONT = "Arial"
NAVY = "1F3864"; LBLUE = "D9E1F2"; GREY = "F2F2F2"; GREEN = "C6EFCE"; AMBER = "FFEB9C"
BLUE_TXT = "0000FF"; thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, ncols, fill=NAVY, color="FFFFFF"):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=color, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, bold=True, size=16, color=NAVY)
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, italic=True, size=10, color="595959")

# ---------------------------------------------------------------- READ ME
ws = wb.active; ws.title = "READ ME"; ws.sheet_view.showGridLines = False
title(ws, "WC2026 xG Value Engine",
      "Deterministic Dixon-Coles / xG model  -  no black box, no LLM in the prediction path")
notes = [
 "", "WHAT THIS IS",
 "A transparent scoreline model. It turns two teams' attack/defence ratings into an expected-goals (xG)",
 "figure for each side, builds the full distribution of scorelines (Dixon-Coles bivariate Poisson), and",
 "prices every market off that one matrix. All maths is in the Python files; nothing is guessed.",
 "",
 "THE WORKFLOW (each runs in seconds)",
 "1. RATINGS tab  - refresh attack/defence from rolling xG (FBref / Understat / FotMob). This is the job.",
 "2. Enter the book odds you see into data/fixtures_odds.csv  (or the Value Finder tab here).",
 "3. Run  python3 value_finder.py  - it flags +EV bets, sized at 1/4 Kelly.",
 "4. Log every bet in the CLV TRACKER tab. Add the closing odds later.",
 "",
 "HOW TO JUDGE YOURSELF  (the only honest scorecard)",
 "Over ~48 remaining games, results are noise. Closing-Line Value is the signal: if you consistently beat",
 "the closing line you have an edge, even on a losing run. If you don't, no winning streak is real.",
 "",
 "WHERE THE EDGE ACTUALLY LIVES  (from the research)",
 "- Small markets: team totals, BTTS, totals. Books shade these least and limit them low. PRIMARY TARGET.",
 "- Danger-zone favourites (odds ~1.34-1.50 / -200 to -299): the public overpays; ~-25% ROI historically.",
 "  The finder TAGS these so you fade, not follow.",
 "- The 48-team format adds mismatches (totals/handicaps) and dead third games (rotation the market lags).",
 "",
 "KNOWN FAILURE MODE  (read this)",
 "Out of the box the ratings are COARSE SEED PRIORS. A coarse model over-rates longshots, so it will flag",
 "things like 'minnow to beat an elite side' as value. That is an artifact, not an edge - those rows are",
 "tagged LONGSHOT - low confidence and sunk to the bottom. Fix it by refreshing ratings before trusting 1X2.",
 "",
 "REALITY CHECK",
 "64 games, ~48 left. Even a real 4% edge can't prove itself before the final. This is a modelling challenge",
 "for curiosity - a clean, honest tool - not a money printer. Stake only what you'd be happy to lose.",
]
for i, line in enumerate(notes, start=4):
    cell = ws.cell(row=i, column=1, value=line)
    if line.isupper() and line.strip():
        cell.font = Font(name=FONT, bold=True, size=11, color=NAVY)
    else:
        cell.font = Font(name=FONT, size=10)
ws.column_dimensions["A"].width = 112

# ---------------------------------------------------------------- RATINGS
rt = pd.read_csv("data/team_ratings.csv")
ws = wb.create_sheet("Ratings"); ws.sheet_view.showGridLines = False
title(ws, "Team Ratings  (xG strength multipliers)",
      "BLUE = edit these. 1.00 = average WC side. atk>1 scores more; dfn<1 concedes fewer. REFRESH FROM ROLLING xG.")
hdr = ["Team", "Attack (atk)", "Defence (dfn)", "Tier", "Notes"]
ws.append([]); ws.append([]); ws.append(hdr)
style_header(ws, 5, len(hdr))
for _, r in rt.iterrows():
    ws.append([r["team"], r["atk"], r["dfn"], r["tier"], r["notes"]])
for row in range(6, 6 + len(rt)):
    for col in (2, 3):
        ws.cell(row=row, column=col).font = Font(name=FONT, color=BLUE_TXT)
        ws.cell(row=row, column=col).number_format = "0.00"
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = BORDER
        if row % 2 == 0:
            if ws.cell(row=row, column=col).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=GREY)
for col, w in zip("ABCDE", (26, 13, 13, 12, 40)):
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- VALUE FINDER
vf = find_value("data/fixtures_odds.csv", "data/team_ratings.csv")
ws = wb.create_sheet("Value Finder"); ws.sheet_view.showGridLines = False
title(ws, "Value Finder  -  latest run",
      "Snapshot from value_finder.py. BLUE = inputs (model prob from engine; offered odds you saw). "
      "Fair odds & edge are live formulas. Re-run the script to refresh.")
hdr = ["Match", "Market", "Selection", "Model prob", "Fair odds",
       "Offered odds", "Edge (EV/unit)", "1/4-Kelly (u, /100)", "Flag"]
ws.append([]); ws.append([]); ws.append(hdr)
style_header(ws, 5, len(hdr))
start = 6
for i, (_, r) in enumerate(vf.iterrows()):
    row = start + i
    ws.cell(row=row, column=1, value=r["match"])
    ws.cell(row=row, column=2, value=r["group"])
    ws.cell(row=row, column=3, value=r["selection"])
    ws.cell(row=row, column=4, value=round(float(r["model_prob"]), 4))
    ws.cell(row=row, column=5, value=f"=IF(D{row}=0,\"\",1/D{row})")
    ws.cell(row=row, column=6, value=round(float(r["offered_odds"]), 3))
    ws.cell(row=row, column=7, value=f"=D{row}*(F{row}-1)-(1-D{row})")
    ws.cell(row=row, column=8, value=float(r["qtr_kelly_units"]))
    ws.cell(row=row, column=9, value=r["flag"])
    ws.cell(row=row, column=4).font = Font(name=FONT, color=BLUE_TXT)
    ws.cell(row=row, column=6).font = Font(name=FONT, color=BLUE_TXT)
    ws.cell(row=row, column=4).number_format = "0.0%"
    ws.cell(row=row, column=5).number_format = "0.00"
    ws.cell(row=row, column=6).number_format = "0.00"
    ws.cell(row=row, column=7).number_format = "0.0%"
    ws.cell(row=row, column=8).number_format = "0.00"
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = BORDER
end = start + len(vf) - 1
ws.conditional_formatting.add(f"G{start}:G{end}",
    CellIsRule(operator="greaterThan", formula=["0.05"],
               fill=PatternFill("solid", fgColor=GREEN)))
ws.conditional_formatting.add(f"I{start}:I{end}",
    CellIsRule(operator="containsText", formula=['"DANGER"'],
               fill=PatternFill("solid", fgColor=AMBER)))
for col, w in zip("ABCDEFGHI", (20, 16, 16, 11, 10, 12, 14, 18, 34)):
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- CLV TRACKER
ws = wb.create_sheet("CLV Tracker"); ws.sheet_view.showGridLines = False
title(ws, "Closing-Line Value Tracker",
      "Log every bet. Add closing odds when the market settles. CLV% and P&L are live formulas. "
      "Result: 1 = win, 0 = loss, 0.5 = push.")
hdr = ["Date", "Match", "Market", "Selection", "Odds taken", "Stake (u)",
       "Closing odds", "Result", "CLV %", "P&L (u)"]
ws.append([]); ws.append([]); ws.append(hdr)
style_header(ws, 5, len(hdr))
first = 6; last = 55
for row in range(first, last + 1):
    ws.cell(row=row, column=9,
            value=f'=IF(OR($E{row}="",$G{row}=""),"",$E{row}/$G{row}-1)')
    ws.cell(row=row, column=10,
            value=(f'=IF(OR($E{row}="",$H{row}=""),"",'
                   f'IF($H{row}=0.5,0,IF($H{row}=1,$F{row}*($E{row}-1),-$F{row})))'))
    ws.cell(row=row, column=9).number_format = "0.0%"
    ws.cell(row=row, column=10).number_format = "0.00;(0.00)"
    for col in (5, 6, 7, 8):
        ws.cell(row=row, column=col).font = Font(name=FONT, color=BLUE_TXT)
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = BORDER

# summary block
s = last + 2
summ = [
    ("Bets logged", f'=COUNT(I{first}:I{last})'),
    ("Avg CLV %", f'=IFERROR(AVERAGE(I{first}:I{last}),0)'),
    ("% beating close", f'=IFERROR(COUNTIF(I{first}:I{last},">0")/COUNT(I{first}:I{last}),0)'),
    ("Units staked", f'=SUM(F{first}:F{last})'),
    ("P&L (units)", f'=SUM(J{first}:J{last})'),
    ("ROI %", f'=IFERROR(SUM(J{first}:J{last})/SUM(F{first}:F{last}),0)'),
]
for i, (lab, fml) in enumerate(summ):
    r = s + i
    ws.cell(row=r, column=2, value=lab).font = Font(name=FONT, bold=True, color=NAVY)
    c = ws.cell(row=r, column=3, value=fml)
    c.font = Font(name=FONT, bold=True)
    c.fill = PatternFill("solid", fgColor=LBLUE)
    if "%" in lab:
        c.number_format = "0.0%"
    elif lab in ("Units staked", "P&L (units)"):
        c.number_format = "0.00;(0.00)"
ws.cell(row=s-1, column=2, value="SUMMARY").font = Font(name=FONT, bold=True, size=12, color=NAVY)
for col, w in zip("ABCDEFGHIJ", (11, 20, 14, 18, 11, 10, 12, 9, 10, 11)):
    ws.column_dimensions[col].width = w

# seed two example rows so the formulas show live
ex = [["2026-06-22", "Brazil v Morocco", "Team Totals", "Home Over 1.5", 2.55, 1.0, 2.30, 1],
      ["2026-06-22", "Neth v Japan", "Totals", "Under 2.5", 2.25, 1.0, 2.05, 0]]
for i, e in enumerate(ex):
    for j, v in enumerate(e, start=1):
        ws.cell(row=first + i, column=j, value=v)

wb.save("WC2026_Value_Engine.xlsx")
print("workbook written:", "WC2026_Value_Engine.xlsx", "| value rows:", len(vf))
